from __future__ import annotations

import csv
import io
import os
import re
import sqlite3
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from flask import Flask, Response, jsonify, request, send_from_directory, session
from openpyxl import load_workbook
from pypdf import PdfReader
from werkzeug.exceptions import HTTPException
from werkzeug.utils import secure_filename


ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "expense.db"
UPLOAD_DIR = ROOT / "uploads" / "invoices"
FINANCE_NAME = "唐磊"
FINANCE_PASSWORD = os.environ.get("FINANCE_PASSWORD", "AT2022")
APP_SECRET_KEY = os.environ.get("APP_SECRET_KEY", "dev-expense-session-key")
DEMO_MODE = os.environ.get("DEMO_MODE", "0") == "1"
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
STATUSES = {"待财务审核", "已驳回", "待付款", "已付款"}
ROLES = {"employee", "finance"}
RECEIPT_TYPES = {"数电发票", "纸质发票", "截图", "收据", "无票据"}
CATEGORIES = [
    "差旅交通",
    "餐饮招待",
    "办公用品",
    "快递物流",
    "签证/证件/行政",
    "软件/订阅服务",
    "业务采购",
    "其他",
]

app = Flask(__name__, static_folder=None)
app.secret_key = APP_SECRET_KEY


def now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def row_to_dict(row: sqlite3.Row | None) -> dict[str, Any] | None:
    return dict(row) if row else None


def mask_bank(account: str | None) -> str:
    digits = "".join(ch for ch in (account or "") if ch.isdigit())
    if len(digits) <= 8:
        return account or ""
    return f"{digits[:4]} **** **** {digits[-4:]}"


def normalize_role(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"财务", "finance", "财务人员"}:
        return "finance"
    if text in {"员工", "employee", "普通员工"}:
        return "employee"
    return "__invalid__"


def effective_role(user: dict[str, Any]) -> str:
    return "finance" if user["name"] == FINANCE_NAME and user["role"] == "finance" else "employee"


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_receipt_type(value: Any) -> str:
    text = normalize_text(value)
    return "数电发票" if text == "电子发票" else text


def claim_total(lines: list[dict[str, Any]]) -> float:
    return round(sum(float(line.get("amount") or 0) for line in lines), 2)


def money_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return round(float(value), 2)


def allowed_upload(filename: str, receipt_type: str) -> bool:
    suffix = Path(filename).suffix.lower()
    if receipt_type == "数电发票":
        return suffix == ".pdf"
    return suffix == ".pdf" or suffix in IMAGE_EXTENSIONS


def validate_upload(file: Any, receipt_type: str) -> str:
    if not file:
        raise ValueError("请上传附件")
    if not file.filename:
        raise ValueError("附件文件名无效")
    filename = secure_filename(file.filename)
    if not allowed_upload(filename, receipt_type):
        if receipt_type == "数电发票":
            raise ValueError("数电发票请上传 PDF 原件")
        raise ValueError("附件仅支持 PDF、JPG、PNG、WEBP")
    file.stream.seek(0, os.SEEK_END)
    size = file.stream.tell()
    file.stream.seek(0)
    if size > MAX_UPLOAD_BYTES:
        raise ValueError("单个附件不能超过 10MB")
    return filename


def parse_invoice_qr_amount(value: str) -> float | None:
    parts = [part.strip() for part in value.split(",")]
    if len(parts) < 5 or parts[0] != "01":
        return None
    amount = parts[4]
    if not re.fullmatch(r"[0-9][0-9,]*\.?[0-9]{0,2}", amount):
        return None
    return round(float(amount.replace(",", "")), 2)


def decode_qr_amount_from_image_bytes(data: bytes) -> float | None:
    os.environ.setdefault("OPENCV_LOG_LEVEL", "SILENT")
    try:
        import cv2
        import numpy as np
    except ImportError:
        return None

    if hasattr(cv2, "setLogLevel"):
        cv2.setLogLevel(0)

    image = cv2.imdecode(np.frombuffer(data, np.uint8), cv2.IMREAD_COLOR)
    if image is None:
        return None

    detector = cv2.QRCodeDetector()
    height, width = image.shape[:2]
    crops = [
        image[: max(1, int(height * 0.35)), : max(1, int(width * 0.35))],
        image,
    ]

    for candidate in crops:
        text, _, _ = detector.detectAndDecode(candidate)
        amount = parse_invoice_qr_amount(text or "")
        if amount is not None:
            return amount
        if hasattr(detector, "detectAndDecodeMulti"):
            ok, decoded, _, _ = detector.detectAndDecodeMulti(candidate)
            if ok:
                for item in decoded:
                    amount = parse_invoice_qr_amount(item or "")
                    if amount is not None:
                        return amount
    return None


def extract_pdf_qr_amount(reader: PdfReader) -> float | None:
    for page in reader.pages:
        for image in getattr(page, "images", []):
            amount = decode_qr_amount_from_image_bytes(image.data)
            if amount is not None:
                return amount
    return None


def extract_pdf_amount(path: Path) -> float | None:
    reader = PdfReader(str(path))
    qr_amount = extract_pdf_qr_amount(reader)
    if qr_amount is not None:
        return qr_amount

    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    if not text.strip():
        return None

    tax_total_index = text.find("价税合计")
    if tax_total_index >= 0:
        segment = text[tax_total_index : tax_total_index + 500]
        amounts = [
            float(match.replace(",", ""))
            for match in re.findall(r"[¥￥]\s*([0-9][0-9,]*\.?[0-9]{0,2})", segment)
        ]
        if amounts:
            return round(max(amounts), 2)

    tax_total_patterns = [
        r"价税合计[\s\S]{0,120}?[（(]\s*小写\s*[）)]\s*[¥￥]?\s*([0-9][0-9,]*\.?[0-9]{0,2})",
        r"价税合计[^\d¥￥]{0,80}[¥￥]\s*([0-9][0-9,]*\.?[0-9]{0,2})",
        r"含税(?:金额|合计|总额)?[^\d¥￥]{0,80}[¥￥]?\s*([0-9][0-9,]*\.?[0-9]{0,2})",
        r"(?:tax\s*total|total\s*with\s*tax|amount\s*with\s*tax)[^\d¥￥]{0,80}[¥￥]?\s*([0-9][0-9,]*\.?[0-9]{0,2})",
    ]
    for pattern in tax_total_patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return round(float(match.group(1).replace(",", "")), 2)

    currency_amounts = [
        float(match.replace(",", ""))
        for match in re.findall(r"[¥￥]\s*([0-9][0-9,]*\.?[0-9]{0,2})", text)
    ]
    if currency_amounts:
        return round(max(currency_amounts), 2)
    return None


def attachment_url(path: str) -> str:
    if not path:
        return ""
    return f"/api/attachments/{Path(path).name}"


def next_claim_id(conn: sqlite3.Connection) -> str:
    ym = datetime.now().strftime("%Y%m")
    prefix = f"BX-{ym}-"
    row = conn.execute(
        "SELECT id FROM expense_claims WHERE id LIKE ? ORDER BY id DESC LIMIT 1",
        (f"{prefix}%",),
    ).fetchone()
    next_num = 1
    if row:
        try:
            next_num = int(row["id"].split("-")[-1]) + 1
        except ValueError:
            next_num = 1
    return f"{prefix}{next_num:03d}"


def get_user(conn: sqlite3.Connection, user_id: int | None = None) -> dict[str, Any]:
    if user_id is None:
        raw = session.get("user_id")
        if not raw:
            raise PermissionError("请先登录")
        user_id = int(raw)
    row = conn.execute("SELECT * FROM users WHERE id = ? AND active = 1", (user_id,)).fetchone()
    if not row:
        raise PermissionError("用户不存在或已停用")
    user = dict(row)
    user["bank_masked"] = mask_bank(user.get("bank_account"))
    return user


def require_finance(conn: sqlite3.Connection) -> dict[str, Any]:
    user = get_user(conn)
    if effective_role(user) != "finance":
        raise PermissionError("需要财务权限")
    return user


def public_user(user: dict[str, Any], include_full_bank: bool = False) -> dict[str, Any]:
    payload = {
        "id": user["id"],
        "name": user["name"],
        "department": user["department"],
        "role": effective_role(user),
        "raw_role": user["role"],
        "bank_name": user["bank_name"],
        "bank_masked": mask_bank(user["bank_account"]),
        "contact": user["contact"],
        "active": bool(user["active"]),
    }
    if include_full_bank:
        payload["bank_account"] = user["bank_account"]
    return payload


def list_claims(conn: sqlite3.Connection, actor: dict[str, Any], scope: str = "auto") -> list[dict[str, Any]]:
    params: list[Any] = []
    where = "1=1"
    actor_role = effective_role(actor)
    if actor_role != "finance" or scope == "mine":
        where = "c.employee_id = ?"
        params.append(actor["id"])

    status = request.args.get("status")
    if status and status != "all":
        where += " AND c.status = ?"
        params.append(status)

    employee_id = request.args.get("employee_id")
    if employee_id and actor_role == "finance":
        where += " AND c.employee_id = ?"
        params.append(int(employee_id))

    department = request.args.get("department")
    if department and actor_role == "finance":
        where += " AND u.department = ?"
        params.append(department)

    query = normalize_text(request.args.get("q")).lower()
    if query:
        where += " AND (lower(u.name) LIKE ? OR lower(u.department) LIKE ? OR lower(c.id) LIKE ? OR lower(c.summary) LIKE ?)"
        like = f"%{query}%"
        params.extend([like, like, like, like])

    month = normalize_text(request.args.get("month"))
    if month:
        where += " AND substr(c.created_at, 1, 7) = ?"
        params.append(month)

    rows = conn.execute(
        f"""
        SELECT c.*, u.name AS employee_name, u.department, u.bank_name, u.bank_account, u.contact
        FROM expense_claims c
        JOIN users u ON u.id = c.employee_id
        WHERE {where}
        ORDER BY c.created_at DESC
        """,
        params,
    ).fetchall()
    claims = []
    for row in rows:
        claim = dict(row)
        line_rows = conn.execute(
            """
            SELECT id, date, category, purpose, amount, receipt_type, attachment_name, attachment_path, invoice_amount, no_receipt_note
            FROM expense_lines
            WHERE claim_id = ?
            ORDER BY id
            """,
            (claim["id"],),
        ).fetchall()
        claim["lines"] = [dict(line) for line in line_rows]
        for line in claim["lines"]:
            line["attachment_url"] = attachment_url(line.get("attachment_path", ""))
            line["issues"] = find_line_issues(line)
        bank_account = claim.pop("bank_account")
        claim["employee"] = {
            "id": claim.pop("employee_id"),
            "name": claim.pop("employee_name"),
            "department": claim.pop("department"),
            "bank_name": claim.pop("bank_name"),
            "bank_account": bank_account if actor_role == "finance" else None,
            "bank_masked": mask_bank(bank_account),
            "contact": claim.pop("contact"),
        }
        claim["issues"] = find_issues(claim)
        claims.append(claim)
    return claims


def find_line_issues(line: dict[str, Any]) -> list[str]:
    issues: list[str] = []
    today = datetime.now().date()
    receipt = normalize_receipt_type(line.get("receipt_type"))
    attachment = normalize_text(line.get("attachment_name"))
    amount = float(line.get("amount") or 0)
    invoice_amount = line.get("invoice_amount")
    if receipt != "无票据" and not attachment:
        issues.append("缺附件")
    if receipt == "数电发票" and invoice_amount is None:
        issues.append("数电发票未识别")
    if receipt == "数电发票" and invoice_amount is not None and abs(float(invoice_amount) - amount) > 0.01:
        issues.append("发票金额不一致")
    if amount >= 1000:
        issues.append("大额")
    if receipt == "无票据":
        issues.append("无票据")
        if not normalize_text(line.get("no_receipt_note")):
            issues.append("缺无票据说明")
    try:
        line_date = datetime.strptime(str(line.get("date")), "%Y-%m-%d").date()
        if (today - line_date).days > 90:
            issues.append("超90天")
        if line_date > today:
            issues.append("未来日期")
    except ValueError:
        issues.append("日期格式异常")
    return issues


def find_issues(claim: dict[str, Any]) -> list[str]:
    issues: list[str] = []
    seen: set[str] = set()
    for line in claim.get("lines", []):
        key = f"{line.get('date')}-{line.get('amount')}-{line.get('purpose')}".lower()
        if key in seen:
            issues.append("疑似重复明细")
        seen.add(key)
        issues.extend(line.get("issues") or find_line_issues(line))
    return list(dict.fromkeys(issues))


def init_db() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    with connect() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              name TEXT NOT NULL,
              department TEXT NOT NULL,
              role TEXT NOT NULL CHECK(role IN ('employee', 'finance')),
              bank_account TEXT NOT NULL DEFAULT '',
              bank_name TEXT NOT NULL DEFAULT '',
              contact TEXT NOT NULL DEFAULT '',
              active INTEGER NOT NULL DEFAULT 1,
              created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
              updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS expense_claims (
              id TEXT PRIMARY KEY,
              employee_id INTEGER NOT NULL REFERENCES users(id),
              summary TEXT NOT NULL DEFAULT '',
              status TEXT NOT NULL CHECK(status IN ('待财务审核', '已驳回', '待付款', '已付款')),
              total_amount REAL NOT NULL,
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL,
              reviewed_by INTEGER REFERENCES users(id),
              review_note TEXT NOT NULL DEFAULT '',
              paid_at TEXT,
              payment_batch_id TEXT REFERENCES payment_batches(id)
            );

            CREATE TABLE IF NOT EXISTS expense_lines (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              claim_id TEXT NOT NULL REFERENCES expense_claims(id) ON DELETE CASCADE,
              date TEXT NOT NULL,
              category TEXT NOT NULL,
              purpose TEXT NOT NULL,
              amount REAL NOT NULL,
              receipt_type TEXT NOT NULL,
              attachment_name TEXT NOT NULL DEFAULT '',
              attachment_path TEXT NOT NULL DEFAULT '',
              invoice_amount REAL,
              no_receipt_note TEXT NOT NULL DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS payment_batches (
              id TEXT PRIMARY KEY,
              created_by INTEGER NOT NULL REFERENCES users(id),
              created_at TEXT NOT NULL,
              total_amount REAL NOT NULL,
              status TEXT NOT NULL,
              exported_at TEXT
            );
            """
        )
        ensure_columns(conn)
        count = conn.execute("SELECT COUNT(*) AS count FROM users").fetchone()["count"]
        if count == 0:
            seed_demo(conn)
        conn.commit()


def ensure_columns(conn: sqlite3.Connection) -> None:
    line_columns = {row["name"] for row in conn.execute("PRAGMA table_info(expense_lines)").fetchall()}
    if "attachment_path" not in line_columns:
        conn.execute("ALTER TABLE expense_lines ADD COLUMN attachment_path TEXT NOT NULL DEFAULT ''")
    if "invoice_amount" not in line_columns:
        conn.execute("ALTER TABLE expense_lines ADD COLUMN invoice_amount REAL")
    if "no_receipt_note" not in line_columns:
        conn.execute("ALTER TABLE expense_lines ADD COLUMN no_receipt_note TEXT NOT NULL DEFAULT ''")


def seed_demo(conn: sqlite3.Connection) -> None:
    users = [
        ("唐磊", "财务部", "finance", "6222000000008812", "招商银行上海分行", "tanglei@example.com"),
        ("诸晨威", "行政", "employee", "6222000000003021", "工商银行上海分行", "zhuchenwei@example.com"),
        ("Andrea Tang", "行政", "employee", "6222000000005168", "中国银行上海分行", "andrea@example.com"),
        ("林一", "销售", "employee", "6228000000009188", "建设银行上海分行", "linyi@example.com"),
    ]
    conn.executemany(
        """
        INSERT INTO users (name, department, role, bank_account, bank_name, contact)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        users,
    )
    claims = [
        (
            "BX-202606-001",
            2,
            "办理 Andrea 工作签相关材料费用。",
            "待财务审核",
            283.34,
            "2026-06-21T10:12:00+08:00",
            "",
        ),
        (
            "BX-202606-002",
            1,
            "税务局领票交通和电子口岸 IC 卡到付邮费。",
            "待财务审核",
            21,
            "2026-06-22T14:31:00+08:00",
            "",
        ),
        (
            "BX-202606-003",
            4,
            "客户拜访交通与餐饮。",
            "待付款",
            704,
            "2026-06-19T09:16:00+08:00",
            "票据齐全，准予付款。",
        ),
    ]
    conn.executemany(
        """
        INSERT INTO expense_claims (id, employee_id, summary, status, total_amount, created_at, updated_at, review_note)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [(id_, uid, summary, status, total, created, now_iso(), note) for id_, uid, summary, status, total, created, note in claims],
    )
    lines = [
        ("BX-202606-001", "2026-05-13", "签证/证件/行政", "办理 Andrea 工作签", 76.34, "数电发票", "工作签-发票-76.34.pdf", "", 76.34),
        ("BX-202606-001", "2026-05-13", "签证/证件/行政", "办理 Andrea 工作签", 87, "数电发票", "工作签-发票-87.pdf", "", 87),
        ("BX-202606-001", "2026-05-13", "签证/证件/行政", "办理 Andrea 工作签", 120, "数电发票", "工作签-发票-120.pdf", "", 120),
        ("BX-202606-002", "2026-06-18", "差旅交通", "去税务局领发票来回地铁费", 10, "纸质发票", "", "", None),
        ("BX-202606-002", "2026-06-20", "快递物流", "电子口岸 IC 卡到付邮费", 11, "数电发票", "IC卡邮费.pdf", "", 11),
        ("BX-202606-003", "2026-06-17", "差旅交通", "拜访华东客户高铁票", 268, "数电发票", "高铁票.pdf", "", 268),
        ("BX-202606-003", "2026-06-17", "餐饮招待", "客户商务餐", 436, "数电发票", "餐饮发票.pdf", "", 486),
    ]
    conn.executemany(
        """
        INSERT INTO expense_lines (claim_id, date, category, purpose, amount, receipt_type, attachment_name, attachment_path, invoice_amount)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        lines,
    )


@app.errorhandler(PermissionError)
def handle_permission(error: PermissionError) -> tuple[Response, int]:
    return jsonify({"error": str(error)}), 403


@app.errorhandler(ValueError)
def handle_value(error: ValueError) -> tuple[Response, int]:
    return jsonify({"error": str(error)}), 400


@app.errorhandler(HTTPException)
def handle_http(error: HTTPException) -> tuple[Response, int]:
    return jsonify({"error": error.description or error.name}), error.code or 500


@app.errorhandler(Exception)
def handle_unexpected(error: Exception) -> tuple[Response, int]:
    app.logger.exception("Unhandled request error")
    return jsonify({"error": "服务器处理失败，请稍后重试或联系财务"}), 500


@app.route("/")
def index() -> Response:
    return send_from_directory(ROOT, "index.html")


@app.get("/assets/<path:filename>")
def asset_files(filename: str) -> Response:
    return send_from_directory(ROOT / "assets", filename)


@app.get("/api/attachments/<path:filename>")
def uploaded_invoice(filename: str) -> Response:
    safe_name = secure_filename(filename)
    if safe_name != filename:
        raise PermissionError("附件不存在或无权访问")
    rel_path = f"uploads/invoices/{safe_name}"
    with connect() as conn:
        actor = get_user(conn)
        row = conn.execute(
            """
            SELECT c.employee_id
            FROM expense_lines l
            JOIN expense_claims c ON c.id = l.claim_id
            WHERE l.attachment_path = ?
            LIMIT 1
            """,
            (rel_path,),
        ).fetchone()
        if not row:
            raise PermissionError("附件不存在或无权访问")
        if effective_role(actor) != "finance" and row["employee_id"] != actor["id"]:
            raise PermissionError("附件不存在或无权访问")
    return send_from_directory(UPLOAD_DIR, safe_name)


@app.route("/<path:path>")
def static_files(path: str) -> Response:
    if path in {"index.html", "app.js", "styles.css"}:
        return send_from_directory(ROOT, path)
    return jsonify({"error": "not found"}), 404


@app.get("/api/login-options")
def login_options() -> Response:
    with connect() as conn:
        rows = conn.execute(
            "SELECT id, name, department, role FROM users WHERE active = 1 ORDER BY department, name"
        ).fetchall()
    users = []
    for row in rows:
        user = dict(row)
        users.append(
            {
                "id": user["id"],
                "name": user["name"],
                "department": user["department"],
                "requires_password": user["name"] == FINANCE_NAME and user["role"] == "finance",
            }
        )
    return jsonify({"users": users})


@app.post("/api/login")
def login() -> Response:
    payload = request.get_json(force=True)
    department = normalize_text(payload.get("department"))
    name = normalize_text(payload.get("name"))
    password = normalize_text(payload.get("password"))
    with connect() as conn:
        row = conn.execute(
            "SELECT * FROM users WHERE active = 1 AND department = ? AND name = ?",
            (department, name),
        ).fetchone()
        if not row:
            raise PermissionError("姓名或部门不匹配")
        actor = dict(row)
        if actor["name"] == FINANCE_NAME and actor["role"] == "finance" and password != FINANCE_PASSWORD:
            raise PermissionError("财务密码不正确")
        session["user_id"] = actor["id"]
        actor["bank_masked"] = mask_bank(actor.get("bank_account"))
        users = [public_user(dict(item), include_full_bank=effective_role(actor) == "finance") for item in conn.execute("SELECT * FROM users WHERE active = 1 ORDER BY department, name")]
        claims = list_claims(conn, actor)
    return jsonify({"current_user": public_user(actor), "users": users, "claims": claims, "categories": CATEGORIES})


@app.post("/api/logout")
def logout() -> Response:
    session.clear()
    return jsonify({"ok": True})


@app.get("/api/bootstrap")
def bootstrap() -> Response:
    with connect() as conn:
        actor = get_user(conn)
        users = [public_user(dict(row), include_full_bank=effective_role(actor) == "finance") for row in conn.execute("SELECT * FROM users WHERE active = 1 ORDER BY department, name")]
        claims = list_claims(conn, actor)
    return jsonify({"current_user": public_user(actor), "users": users, "claims": claims, "categories": CATEGORIES})


@app.post("/api/invoices/inspect")
def inspect_invoice() -> Response:
    with connect() as conn:
        get_user(conn)
    file = request.files.get("file")
    receipt_type = normalize_receipt_type(request.form.get("receipt_type"))
    if receipt_type not in RECEIPT_TYPES or receipt_type == "无票据":
        raise ValueError("票据类型无效")
    filename = validate_upload(file, receipt_type)
    stored_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8]}-{filename}"
    path = UPLOAD_DIR / stored_name
    file.save(path)
    amount = None
    if receipt_type == "数电发票":
        amount = extract_pdf_amount(path)
    attachment_path = f"uploads/invoices/{stored_name}"
    return jsonify(
        {
            "attachment_name": filename,
            "attachment_path": attachment_path,
            "attachment_url": attachment_url(attachment_path),
            "invoice_amount": amount,
        }
    )


@app.get("/api/users")
def users() -> Response:
    with connect() as conn:
        actor = require_finance(conn)
        rows = conn.execute("SELECT * FROM users ORDER BY active DESC, role DESC, name").fetchall()
    return jsonify({"users": [public_user(dict(row), include_full_bank=True) for row in rows]})


@app.post("/api/users")
def create_or_update_user() -> Response:
    payload = request.get_json(force=True)
    with connect() as conn:
        require_finance(conn)
        user = {
            "name": normalize_text(payload.get("name")),
            "department": normalize_text(payload.get("department")),
            "role": normalize_role(payload.get("role")),
            "bank_account": normalize_text(payload.get("bank_account")),
            "bank_name": normalize_text(payload.get("bank_name")),
            "contact": normalize_text(payload.get("contact")),
        }
        errors = validate_single_user(user)
        if errors:
            return jsonify({"errors": errors}), 400
        timestamp = now_iso()
        existing = conn.execute("SELECT id FROM users WHERE name = ? ORDER BY active DESC, id DESC LIMIT 1", (user["name"],)).fetchone()
        if existing:
            conn.execute(
                """
                UPDATE users
                SET department = ?, role = ?, bank_account = ?, bank_name = ?, contact = ?, active = 1, updated_at = ?
                WHERE id = ?
                """,
                (
                    user["department"],
                    user["role"],
                    user["bank_account"],
                    user["bank_name"],
                    user["contact"],
                    timestamp,
                    existing["id"],
                ),
            )
        else:
            conn.execute(
                """
                INSERT INTO users (name, department, role, bank_account, bank_name, contact, active, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, 1, ?)
                """,
                (
                    user["name"],
                    user["department"],
                    user["role"],
                    user["bank_account"],
                    user["bank_name"],
                    user["contact"],
                    timestamp,
                ),
            )
        conn.commit()
        rows = conn.execute("SELECT * FROM users ORDER BY active DESC, role DESC, name").fetchall()
    return jsonify({"users": [public_user(dict(row), include_full_bank=True) for row in rows]})


@app.delete("/api/users/<int:user_id>")
def delete_user(user_id: int) -> Response:
    with connect() as conn:
        actor = require_finance(conn)
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            raise ValueError("人员不存在")
        user = dict(row)
        if user["id"] == actor["id"]:
            raise ValueError("不能删除当前登录人员")
        if user["name"] == FINANCE_NAME:
            raise ValueError("不能删除财务负责人")
        conn.execute("UPDATE users SET active = 0, updated_at = ? WHERE id = ?", (now_iso(), user_id))
        conn.commit()
        rows = conn.execute("SELECT * FROM users ORDER BY active DESC, role DESC, name").fetchall()
    return jsonify({"users": [public_user(dict(row), include_full_bank=True) for row in rows], "deleted": user_id})


@app.post("/api/users/import/preview")
def import_users_preview() -> Response:
    with connect() as conn:
        require_finance(conn)
        file = request.files.get("file")
        if not file:
            raise ValueError("请上传 Excel 文件")
        rows, errors = parse_user_workbook(file.read())
        summary = import_summary(conn, rows)
    return jsonify({"rows": rows, "errors": errors, "summary": summary})


@app.post("/api/users/import/confirm")
def import_users_confirm() -> Response:
    payload = request.get_json(force=True)
    rows = payload.get("rows") or []
    if not rows:
        raise ValueError("没有可导入的人员记录")
    errors = validate_user_rows(rows)
    if errors:
        return jsonify({"errors": errors, "imported": 0}), 400

    with connect() as conn:
        require_finance(conn)
        conn.execute("UPDATE users SET active = 0, updated_at = ?", (now_iso(),))
        for row in rows:
            conn.execute(
                """
                INSERT INTO users (name, department, role, bank_account, bank_name, contact, active, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, 1, ?)
                """,
                (
                    row["name"],
                    row["department"],
                    row["role"],
                    row["bank_account"],
                    row["bank_name"],
                    row["contact"],
                    now_iso(),
                ),
            )
        conn.commit()
        users = [public_user(dict(row), include_full_bank=True) for row in conn.execute("SELECT * FROM users WHERE active = 1 ORDER BY role DESC, name")]
    return jsonify({"users": users, "imported": len(rows)})


def parse_user_workbook(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["Excel 为空"]
    header = [normalize_text(value) for value in rows[0]]
    aliases = {
        "姓名": "name",
        "部门": "department",
        "角色": "role",
        "银行卡号": "bank_account",
        "开户行": "bank_name",
        "手机号/邮箱": "contact",
        "联系方式": "contact",
        "邮箱": "contact",
        "手机号": "contact",
    }
    indexes = {aliases[name]: index for index, name in enumerate(header) if name in aliases}
    required = ["name", "department", "role", "bank_account", "bank_name"]
    missing = [field for field in required if field not in indexes]
    if missing:
        return [], [f"缺少字段：{', '.join(missing)}"]

    parsed = []
    for number, values in enumerate(rows[1:], start=2):
        if not any(values):
            continue
        row = {
            "row": number,
            "name": normalize_text(values[indexes["name"]]),
            "department": normalize_text(values[indexes["department"]]),
            "role": normalize_role(values[indexes["role"]]),
            "bank_account": normalize_text(values[indexes["bank_account"]]),
            "bank_name": normalize_text(values[indexes["bank_name"]]),
            "contact": normalize_text(values[indexes.get("contact", -1)]) if "contact" in indexes else "",
        }
        parsed.append(row)
    return parsed, validate_user_rows(parsed)


def validate_single_user(row: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    name = normalize_text(row.get("name"))
    role = normalize_role(row.get("role"))
    if not name:
        errors.append("缺少姓名")
    if not normalize_text(row.get("department")):
        errors.append("缺少部门")
    if role not in ROLES:
        errors.append("角色必须是 employee 或 finance")
    if name == FINANCE_NAME and role != "finance":
        errors.append("唐磊必须设置为财务角色")
    if role == "finance" and name != FINANCE_NAME:
        errors.append("只有唐磊可以设置为财务角色")
    if not normalize_text(row.get("bank_account")):
        errors.append("缺少银行卡号")
    if not normalize_text(row.get("bank_name")):
        errors.append("缺少开户行")
    return errors


def validate_user_rows(rows: list[dict[str, Any]]) -> list[str]:
    errors: list[str] = []
    names: set[str] = set()
    finance_count = 0
    for index, row in enumerate(rows, start=1):
        label = f"第 {row.get('row', index)} 行"
        name = normalize_text(row.get("name"))
        role = normalize_role(row.get("role"))
        if not name:
            errors.append(f"{label} 缺少姓名")
        if name in names:
            errors.append(f"{label} 姓名重复：{name}")
        names.add(name)
        if not normalize_text(row.get("department")):
            errors.append(f"{label} 缺少部门")
        if role not in ROLES:
            errors.append(f"{label} 角色必须是 employee 或 finance")
        if role == "finance":
            finance_count += 1
            if name != FINANCE_NAME:
                errors.append(f"{label} 只有唐磊可以设置为财务角色")
        if not normalize_text(row.get("bank_account")):
            errors.append(f"{label} 缺少银行卡号")
        if not normalize_text(row.get("bank_name")):
            errors.append(f"{label} 缺少开户行")
    if rows and finance_count == 0:
        errors.append("至少需要一名财务人员")
    return errors


def import_summary(conn: sqlite3.Connection, rows: list[dict[str, Any]]) -> dict[str, Any]:
    existing = {row["name"] for row in conn.execute("SELECT name FROM users WHERE active = 1").fetchall()}
    incoming = {normalize_text(row.get("name")) for row in rows if normalize_text(row.get("name"))}
    finance_names = [normalize_text(row.get("name")) for row in rows if normalize_role(row.get("role")) == "finance"]
    return {
        "total": len(rows),
        "new_count": len(incoming - existing),
        "disabled_count": len(existing - incoming),
        "missing_bank_count": sum(1 for row in rows if not normalize_text(row.get("bank_account"))),
        "finance_ok": finance_names == [FINANCE_NAME],
    }


def normalize_claim_lines(lines: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not lines:
        raise ValueError("至少需要一条费用明细")
    normalized_lines = []
    for line in lines:
        amount = float(line.get("amount") or 0)
        if amount <= 0:
            raise ValueError("报销金额必须大于 0")
        category = normalize_text(line.get("category"))
        if category not in CATEGORIES:
            raise ValueError("费用类型无效")
        date_value = normalize_text(line.get("date"))
        purpose = normalize_text(line.get("purpose"))
        if not date_value or not purpose:
            raise ValueError("请补齐明细日期和用途")
        receipt_type = normalize_receipt_type(line.get("receipt_type"))
        if receipt_type not in RECEIPT_TYPES:
            raise ValueError("票据类型无效")
        attachment_name = normalize_text(line.get("attachment_name"))
        attachment_path = normalize_text(line.get("attachment_path"))
        no_receipt_note = normalize_text(line.get("no_receipt_note"))
        if receipt_type == "无票据":
            attachment_name = ""
            attachment_path = ""
            if not no_receipt_note:
                raise ValueError("无票据明细必须填写说明")
        else:
            if not attachment_name or not attachment_path:
                raise ValueError("每条非无票据明细都必须上传附件")
            if not allowed_upload(attachment_name, receipt_type):
                if receipt_type == "数电发票":
                    raise ValueError("数电发票请上传 PDF 原件")
                raise ValueError("附件仅支持 PDF、JPG、PNG、WEBP")
        normalized_lines.append(
            {
                "date": date_value,
                "category": category,
                "purpose": purpose,
                "amount": amount,
                "receipt_type": receipt_type,
                "attachment_name": attachment_name,
                "attachment_path": attachment_path,
                "invoice_amount": money_or_none(line.get("invoice_amount")) if receipt_type == "数电发票" else None,
                "no_receipt_note": no_receipt_note,
            }
        )
    return normalized_lines


def validate_line_attachments(
    conn: sqlite3.Connection,
    actor: dict[str, Any],
    lines: list[dict[str, Any]],
    current_claim_id: str | None = None,
) -> None:
    for line in lines:
        attachment_path = normalize_text(line.get("attachment_path"))
        if not attachment_path:
            continue
        row = conn.execute(
            """
            SELECT c.id, c.employee_id
            FROM expense_lines l
            JOIN expense_claims c ON c.id = l.claim_id
            WHERE l.attachment_path = ?
            LIMIT 1
            """,
            (attachment_path,),
        ).fetchone()
        if row and (row["employee_id"] != actor["id"] or (current_claim_id and row["id"] != current_claim_id)):
            raise PermissionError("附件不存在或无权使用")


@app.get("/api/claims")
def claims() -> Response:
    with connect() as conn:
        actor = get_user(conn)
        data = list_claims(conn, actor, request.args.get("scope", "auto"))
    return jsonify({"claims": data})


@app.post("/api/claims")
def create_claim() -> Response:
    payload = request.get_json(force=True)
    with connect() as conn:
        actor = get_user(conn)
        if not actor["bank_account"] or not actor["bank_name"]:
            raise ValueError("银行卡信息缺失，请联系财务维护")
        normalized_lines = normalize_claim_lines(payload.get("lines") or [])
        validate_line_attachments(conn, actor, normalized_lines)
        claim_id = next_claim_id(conn)
        total = claim_total(normalized_lines)
        timestamp = now_iso()
        conn.execute(
            """
            INSERT INTO expense_claims (id, employee_id, summary, status, total_amount, created_at, updated_at)
            VALUES (?, ?, ?, '待财务审核', ?, ?, ?)
            """,
            (claim_id, actor["id"], normalize_text(payload.get("summary")), total, timestamp, timestamp),
        )
        conn.executemany(
            """
            INSERT INTO expense_lines (claim_id, date, category, purpose, amount, receipt_type, attachment_name, attachment_path, invoice_amount, no_receipt_note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    claim_id,
                    line["date"],
                    line["category"],
                    line["purpose"],
                    line["amount"],
                    line["receipt_type"],
                    line["attachment_name"],
                    line["attachment_path"],
                    line["invoice_amount"],
                    line["no_receipt_note"],
                )
                for line in normalized_lines
            ],
        )
        conn.commit()
        claims = list_claims(conn, actor)
    return jsonify({"claim": next((claim for claim in claims if claim["id"] == claim_id), None)})


@app.patch("/api/claims/<claim_id>/supplement")
def supplement_claim(claim_id: str) -> Response:
    payload = request.get_json(force=True)
    with connect() as conn:
        actor = get_user(conn)
        row = conn.execute("SELECT * FROM expense_claims WHERE id = ?", (claim_id,)).fetchone()
        if not row:
            raise ValueError("报销单不存在")
        if row["employee_id"] != actor["id"]:
            raise PermissionError("只能补充自己的报销单")
        if row["status"] != "已驳回":
            raise ValueError("只有已驳回状态可以重新提交")
        if not actor["bank_account"] or not actor["bank_name"]:
            raise ValueError("银行卡信息缺失，请联系财务维护")
        normalized_lines = normalize_claim_lines(payload.get("lines") or [])
        validate_line_attachments(conn, actor, normalized_lines, claim_id)
        total = claim_total(normalized_lines)
        timestamp = now_iso()
        conn.execute("DELETE FROM expense_lines WHERE claim_id = ?", (claim_id,))
        conn.executemany(
            """
            INSERT INTO expense_lines (claim_id, date, category, purpose, amount, receipt_type, attachment_name, attachment_path, invoice_amount, no_receipt_note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    claim_id,
                    line["date"],
                    line["category"],
                    line["purpose"],
                    line["amount"],
                    line["receipt_type"],
                    line["attachment_name"],
                    line["attachment_path"],
                    line["invoice_amount"],
                    line["no_receipt_note"],
                )
                for line in normalized_lines
            ],
        )
        conn.execute(
            """
            UPDATE expense_claims
            SET summary = ?, status = '待财务审核', total_amount = ?, updated_at = ?
            WHERE id = ?
            """,
            (normalize_text(payload.get("summary")), total, timestamp, claim_id),
        )
        conn.commit()
        claims = list_claims(conn, actor)
    return jsonify({"claim": next((claim for claim in claims if claim["id"] == claim_id), None)})


@app.delete("/api/claims/<claim_id>")
def delete_claim(claim_id: str) -> Response:
    with connect() as conn:
        actor = get_user(conn)
        row = conn.execute("SELECT * FROM expense_claims WHERE id = ?", (claim_id,)).fetchone()
        if not row:
            raise ValueError("报销单不存在")
        if row["employee_id"] != actor["id"]:
            raise PermissionError("只能删除自己的报销单")
        if row["status"] != "已驳回":
            raise ValueError("只有已驳回的报销单可以删除")
        conn.execute("DELETE FROM expense_lines WHERE claim_id = ?", (claim_id,))
        conn.execute("DELETE FROM expense_claims WHERE id = ?", (claim_id,))
        conn.commit()
    return jsonify({"deleted": claim_id})


@app.patch("/api/claims/<claim_id>/review")
def review_claim(claim_id: str) -> Response:
    payload = request.get_json(force=True)
    status = normalize_text(payload.get("status"))
    if status not in {"待付款", "已驳回"}:
        raise ValueError("审核状态无效")
    review_note = normalize_text(payload.get("review_note"))
    if status == "已驳回" and not review_note:
        raise ValueError("驳回时必须填写审核备注")
    with connect() as conn:
        actor = require_finance(conn)
        row = conn.execute("SELECT * FROM expense_claims WHERE id = ?", (claim_id,)).fetchone()
        if not row:
            raise ValueError("报销单不存在")
        if row["status"] == "已付款":
            raise ValueError("已付款报销单不能重新审核")
        conn.execute(
            """
            UPDATE expense_claims
            SET status = ?, reviewed_by = ?, review_note = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, actor["id"], review_note, now_iso(), claim_id),
        )
        conn.commit()
        claims = list_claims(conn, actor)
    return jsonify({"claim": next((claim for claim in claims if claim["id"] == claim_id), None)})


@app.post("/api/payments/export")
def export_payments() -> Response:
    payload = request.get_json(force=True)
    ids = payload.get("ids") or []
    if not ids:
        raise ValueError("请选择付款记录")
    with connect() as conn:
        actor = require_finance(conn)
        placeholders = ",".join("?" for _ in ids)
        rows = conn.execute(
            f"""
            SELECT c.id, c.total_amount, c.summary, u.name, u.bank_name, u.bank_account
            FROM expense_claims c
            JOIN users u ON u.id = c.employee_id
            WHERE c.id IN ({placeholders}) AND c.status = '待付款' AND c.payment_batch_id IS NULL
            ORDER BY c.created_at
            """,
            ids,
        ).fetchall()
        if not rows:
            raise ValueError("所选记录中没有待付款报销单")
        batch_id = f"PAY-{datetime.now().strftime('%Y%m%d')}-{datetime.now().strftime('%H%M%S')}"
        total = round(sum(float(row["total_amount"]) for row in rows), 2)
        conn.execute(
            """
            INSERT INTO payment_batches (id, created_by, created_at, total_amount, status, exported_at)
            VALUES (?, ?, ?, ?, '已导出', ?)
            """,
            (batch_id, actor["id"], now_iso(), total, now_iso()),
        )
        conn.execute(
            f"""
            UPDATE expense_claims
            SET payment_batch_id = ?, updated_at = ?
            WHERE id IN ({placeholders}) AND status = '待付款' AND payment_batch_id IS NULL
            """,
            [batch_id, now_iso(), *ids],
        )
        conn.commit()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["报销单号", "收款人", "开户行", "银行卡号", "金额", "付款备注"])
    for row in rows:
        writer.writerow([row["id"], row["name"], row["bank_name"], row["bank_account"], f"{row['total_amount']:.2f}", row["summary"]])
    return Response(
        "\ufeff" + output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=payment-{batch_id}.csv"},
    )


@app.post("/api/payments/mark-paid")
def mark_paid() -> Response:
    payload = request.get_json(force=True)
    ids = payload.get("ids") or []
    if not ids:
        raise ValueError("请选择付款记录")
    with connect() as conn:
        actor = require_finance(conn)
        placeholders = ",".join("?" for _ in ids)
        rows = conn.execute(
            f"SELECT id, total_amount, payment_batch_id FROM expense_claims WHERE id IN ({placeholders}) AND status = '待付款'",
            ids,
        ).fetchall()
        if not rows:
            raise ValueError("没有可标记为已付款的记录")
        existing_batches = {row["payment_batch_id"] for row in rows if row["payment_batch_id"]}
        all_exported_same_batch = len(existing_batches) == 1 and all(row["payment_batch_id"] for row in rows)
        batch_id = next(iter(existing_batches)) if all_exported_same_batch else f"PAY-{datetime.now().strftime('%Y%m%d')}-{datetime.now().strftime('%H%M%S')}"
        total = round(sum(float(row["total_amount"]) for row in rows), 2)
        if not all_exported_same_batch:
            conn.execute(
                """
                INSERT INTO payment_batches (id, created_by, created_at, total_amount, status, exported_at)
                VALUES (?, ?, ?, ?, '已付款', ?)
                """,
                (batch_id, actor["id"], now_iso(), total, now_iso()),
            )
        else:
            conn.execute(
                "UPDATE payment_batches SET status = '已付款' WHERE id = ?",
                (batch_id,),
            )
        conn.execute(
            f"""
            UPDATE expense_claims
            SET status = '已付款', paid_at = ?, payment_batch_id = ?, updated_at = ?
            WHERE id IN ({placeholders}) AND status = '待付款'
            """,
            [now_iso(), batch_id, now_iso(), *ids],
        )
        conn.commit()
        claims = list_claims(conn, actor)
    return jsonify({"batch_id": batch_id, "paid": len(rows), "claims": claims})


@app.get("/api/ledger/export")
def export_ledger() -> Response:
    with connect() as conn:
        actor = require_finance(conn)
        claims = list_claims(conn, actor)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["单号", "提交时间", "报销人", "部门", "费用类型", "用途", "票据类型", "报销金额", "发票金额", "状态", "付款批次", "审核备注", "附件数量"])
    for claim in claims:
        attachment_count = sum(1 for line in claim["lines"] if line.get("attachment_name"))
        for line in claim["lines"]:
            writer.writerow(
                [
                    claim["id"],
                    claim["created_at"],
                    claim["employee"]["name"],
                    claim["employee"]["department"],
                    line["category"],
                    line["purpose"],
                    line["receipt_type"],
                    f"{float(line['amount']):.2f}",
                    "" if line.get("invoice_amount") is None else f"{float(line['invoice_amount']):.2f}",
                    claim["status"],
                    claim.get("payment_batch_id") or "",
                    claim.get("review_note") or "",
                    attachment_count,
                ]
            )
    return Response(
        "\ufeff" + output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=ledger-{datetime.now().strftime('%Y%m%d')}.csv"},
    )


@app.post("/api/reset-demo")
def reset_demo() -> Response:
    payload = request.get_json(force=True)
    with connect() as conn:
        require_finance(conn)
        if not DEMO_MODE and normalize_text(payload.get("confirm")) != "RESET":
            raise ValueError("请确认恢复演示数据")
        conn.executescript(
            """
            DELETE FROM expense_lines;
            DELETE FROM expense_claims;
            DELETE FROM payment_batches;
            DELETE FROM users;
            DELETE FROM sqlite_sequence WHERE name IN ('users', 'expense_lines');
            """
        )
        seed_demo(conn)
        conn.commit()
        actor = get_user(conn, 1)
        session["user_id"] = actor["id"]
        users = [public_user(dict(row), include_full_bank=True) for row in conn.execute("SELECT * FROM users WHERE active = 1 ORDER BY role DESC, name")]
        claims = list_claims(conn, actor)
    return jsonify({"current_user": public_user(actor), "users": users, "claims": claims})


@app.post("/api/clear-claims")
def clear_claims() -> Response:
    with connect() as conn:
        actor = require_finance(conn)
        conn.executescript(
            """
            DELETE FROM expense_lines;
            DELETE FROM expense_claims;
            DELETE FROM payment_batches;
            DELETE FROM sqlite_sequence WHERE name = 'expense_lines';
            """
        )
        conn.commit()
        users = [public_user(dict(row), include_full_bank=True) for row in conn.execute("SELECT * FROM users WHERE active = 1 ORDER BY department, name")]
    return jsonify({"current_user": public_user(actor), "users": users, "claims": [], "cleared": True})


init_db()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5174"))
    app.run(host="127.0.0.1", port=port, debug=True)
